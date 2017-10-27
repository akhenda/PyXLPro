import os
import wget
import multiprocessing as mp
from datetime import datetime
from openpyxl import load_workbook


class PyXLPro(object):
    def __init__(self):
        self.xl = None
        self.xl_name = ''
        self.start_row = 2
        # self.max_rows = ws.max_row + 1
        self.max_rows = 22
        self.url = 'https://forms.iebc.or.ke/'
        self.file_path = 'output/'

    def load_xl(self):
        for file in os.listdir('input'):
            file_path = 'input/' + file
            # print(file)
            self.xl_name = file
            print('\nLoading all Gazetted Polling Stations...')
            self.xl = load_workbook(file_path)
            print('All Gazetted Polling Stations Loaded')
            # print(self.xl.get_sheet_names())
            # self.process_xl(start_row, max_rows)

    def process_xl(self, process, total_jobs):
        print('\n\nWorker No.: %s out of %s' % (process, total_jobs))
        rows_to_process = self.max_rows // total_jobs
        max_rows = rows_to_process * process + 2
        start_row = (process * rows_to_process) - rows_to_process + 2
        print('Start Row: %s' % start_row)
        print('Max Rows: %s' % max_rows)
        ws = self.xl.active
        for row in range(start_row, max_rows):
            print('\n\n\nConstructing File URL...')
            self.construct_url(str(row), ws)
            print('\nConstructing File Output Location...')
            self.construct_file_path(str(row), ws)
            print('\nFetching File...')
            self.fetch_results()

    def construct_url(self, row, worksheet):
        county = '1_' + worksheet['A' + row].value
        constituency = worksheet['C' + row].value
        ward = worksheet['E' + row].value
        polling_centre = worksheet['G' + row].value
        polling_station = worksheet['J' + row].value[-2:]
        self.url += 'storage/f34a/1_'
        self.url += county + '_' + constituency + '_' + ward + '_'
        self.url += polling_centre + '_' + polling_station + '.jpeg'
        print('File URL: %s' % self.url)

    def construct_file_path(self, row, worksheet):
        county = worksheet['B' + row].value
        constituency = worksheet['D' + row].value
        ward = worksheet['F' + row].value
        polling_centre = worksheet['K' + row].value
        self.file_path += county + '/' + constituency + '/' + ward + '/'
        self.file_path += polling_centre + '/'
        print('Output Location: %s' % self.file_path)

    def fetch_results(self):
        if not os.path.exists(self.file_path):
            os.makedirs(self.file_path)
        wget.download(self.url, out=self.file_path)
        self.url = 'https://forms.iebc.or.ke/'
        self.file_path = 'output/'

    def save_xl(self):
        if not os.path.exists('output'):
            os.makedirs('output')
        output_path = 'output/' + self.xl_name
        self.xl.save(output_path)


if __name__ == '__main__':
    start_time = datetime.now()
    run = PyXLPro()
    run.load_xl()
    jobs = []
    jobs_list = range(1, 5)
    for i in jobs_list:
        p = mp.Process(target=run.process_xl(i, len(jobs_list)))
        jobs.append(p)
        p.start()
    print("\n\nThis script executed in %s" % str(datetime.now() - start_time))
